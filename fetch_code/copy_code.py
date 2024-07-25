import pandas as pd
from pandas import DataFrame
import pyperclip as pc
from pathlib import Path
import webbrowser
import openpyxl
from openpyxl import Workbook
import os


def check_index(index: str, programs: dict, prompt: str) -> str:
    """
        Check if the index is valid and return the program name
    """
    try:
        index = int(index)
        return programs[index]

    except KeyError:
        print('Invalid input. Please enter a number from the list.')
        index = int(input(prompt))
        return check_index(index, programs)

    except ValueError:
        print('Invalid input. Please enter a number from the list.')
        index = int(input(prompt))
        return check_index(index, programs)


def get_link_with_code(code, codes_links_df) -> str:
    """
        Get the link to the tool with the project code
    """
    return codes_links_df[codes_links_df['Project code'] == code]['Tool link'].values[0]


def check_stop() -> bool:
    """
        Check if the user wants to stop copying codes
    """
    stop = input('Next? (Y/n): ')
    return stop.lower() == 'n'


def get_last_code(location) -> str:
    """
        Get the last copied code from the last_code.txt file
    """
    try:
        with open(location / 'last_code.txt', 'r') as f:
            last_code = f.read()
    except FileNotFoundError:
        last_code = None
    return last_code


def save_last_code(code, location) -> bool:
    """
        Save the last copied code to the last_code.txt file
    """
    try:
        with open(location / 'last_code.txt', 'w') as f:
            f.write(code)
        return True

    except FileNotFoundError:
        return False

    except Exception as e:
        print(e)
        return False


def edit_imported_indicator(text: str, project_code: str, projects: str) -> bool:
    """
        Edit the excel file to indicate if the indicator was imported
    """
    try:
        wb: Workbook = openpyxl.load_workbook(projects)
        ws = wb.active
        project_row = None

        for i, cell in enumerate(ws['A']):
            if cell.value == project_code:
                project_row = i + 1
                break

        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "imported indicator?":
                    ws.cell(row=project_row, column=cell.column, value=text)
        wb.save(projects)
        wb.close()
        return True

    except FileNotFoundError:
        return False

    except Exception as e:
        print(e)
        return False


def copy_code(code: str, codes_links_df: DataFrame, projects: str, location: Path) -> None:
    if "2019" in code:
        print('Skipping 2019 code...')
        edit_imported_indicator("No, 2019 file", code, projects)
        return
    pc.copy(code)
    print(f'Copied code: {code}')
    save_last_code(code, location)
    file = codes_links_df[codes_links_df['Project code']
                          == code]['file'].values[0]
    indicator = codes_links_df[codes_links_df['Project code']
                               == code]['Indicators in tool'].values[0]
    webbrowser.open(get_link_with_code(code, codes_links_df))
    if str(file) == "nan":
        print('No file found for this code.')
        if str(indicator) == "Yes":
            print("However, the indicators are available in the tool.")
        else:
            return
    else:
        print(f'Found file: {file}')
        open_file = input('Open file? (Y/n): ')
        if open_file.lower() == 'n':
            pass
        else:
            os.startfile(
                codes_links_df[codes_links_df['Project code'] == code]['file'].values[0])
    uploaded_indicator = input('Uploaded indicator? (y/n/e): ') or ''
    if uploaded_indicator.lower() == 'y':
        edit_imported_indicator("Yes", code, projects)
    elif uploaded_indicator.lower() == 'n':
        edit_imported_indicator("No", code, projects)
    elif uploaded_indicator.lower() == 'e':
        text = '\n'.join(line for line in iter(input, ''))
        edit_imported_indicator(text, code, projects)
    else:
        pass


def main(codes: list, location: Path, projects: str, codes_links_df: DataFrame) -> None:
    last_code = get_last_code(location)
    if last_code and last_code in codes:
        start_index = codes.index(last_code) + 1
    else:
        start_index = 0

    for code in codes[start_index:]:
        copy_code(code, codes_links_df, projects, location)
        if check_stop():
            save_last_code(code, location)
            exit()

    save_last_code("", location)
    print('All codes copied!')


# Locations
projects: Path = Path(
    r"C:\Users\DavyBellens\VLIR-UOS\Digital Factory - APR-indicatoren 2022-2023\Lijst projecten.xlsx")

location: Path = Path(__file__).parent

# Read the projects file
try:
    csv: DataFrame = pd.read_excel(projects)
    headers: dict = {i: j for i, j in enumerate(csv.columns, 1)}

    print('Available columns:')
    for i in headers.keys():
        print(f'{i}: {headers[i]}')

    # Select the column with the programs
    prompt: str = "\n\nEnter the number of the column with the programs: "
    program_column: int = int(input(prompt))
    option: str = check_index(program_column, headers, prompt)

    options: dict = {i: j for i, j in enumerate(csv[option].unique(), 1)}

    # Show available options
    print('\nAvailable options:')
    for i in options.keys():
        print(f'{i}: {options[i]}')

    # Select a program
    prompt: str = "\n\nEnter the number of the program you want to copy codes from: "
    program_index: int = int(input(prompt))

    # Get the codes for the selected program
    program: str = check_index(program_index, options, prompt)
    codes_links_df: DataFrame = csv[csv[option]
                                    == program][['Project code', 'Tool link', "file", "Indicators in tool"]]

    codes: list = list(codes_links_df['Project code'])

    # Show the selected program
    print(f'\nProgram selected: {program}')

    main(codes, location, projects, codes_links_df)
    input('\nPress enter to exit.')
    exit()
except PermissionError:
    print('\nPlease close the excel file before running the script.')
    input('\nPress enter to exit.')
    exit()

except Exception as e:
    print(f"\n{e}\n")
    input('\nPress enter to exit.')
    exit()
