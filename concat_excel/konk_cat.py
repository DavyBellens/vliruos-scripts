import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook

projects = Path(
    r'C:\Users\DavyBellens\VLIR-UOS\Digital Factory - APR-indicatoren 2022-2023\Lijst projecten.xlsx')

projects_df = pd.read_excel(projects)

options = "Unable to find match for row 10", "Unable to find match for row 11", "Unable to find match for row 13"

files = list(projects_df[
    projects_df['imported indicator?'].isin(options)
]['file'])

# Loop through each file in the directory
for index, filename in enumerate(files):
    try:
        if str(filename) != "nan":
            print(filename)
            if filename.endswith(".xlsx"):
                # Read the Excel file
                wb = load_workbook(filename)
                ws = wb["NP"]
                # Process rows 10 to 49
                for row in range(10, 50):
                    cell_A = ws[f'A{row}'].value
                    cell_F = ws[f'F{row}'].value

                    cell_str_a = str(cell_A) if cell_A else ""

                    if cell_str_a:
                        cell_str_a = cell_str_a.split(" ")[0]
                        cell_str_a = cell_str_a.split("-")
                        cell_str_a = f"{
                            cell_str_a[1]}/{cell_str_a[2]}/{cell_str_a[0]}"
                    cell_str_f = str(cell_F) if cell_F else ""

                    if cell_str_a and cell_str_f:
                        result = f'Start date: {cell_str_a}, {cell_str_f}'
                    elif cell_str_a and not cell_str_f:
                        result = f"Start date: {cell_str_a}"
                    elif cell_str_f and not cell_str_a:
                        result = cell_str_f
                    else:
                        result = ""

                    # Concatenate A and F columns
                    ws[f'F{row}'] = result if result != "NoneNone" else ""
                    ws[f'A{row}'] = ""

                # Save the changes to the workbook
                wb.save(filename)
                print(f"Processed {filename}")
    except Exception as e:
        print(f"Error processing {filename}: {e}")

print("Files have been processed.")
